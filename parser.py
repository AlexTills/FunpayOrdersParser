from bs4 import BeautifulSoup
import openpyxl
from openpyxl import Workbook
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment
from openpyxl.worksheet.hyperlink import Hyperlink
from colorama import init, Fore, Back, Style
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from selenium.common.exceptions import NoSuchElementException
from tqdm import tqdm
from selenium.webdriver.chrome.options import Options
import os

script_dir = os.path.dirname(os.path.abspath(__file__))
cookie_file_path = os.path.join(script_dir, "cookies.txt")

text = """
███████╗██╗░░░██╗███╗░░██╗██████╗░░█████╗░██╗░░░██╗
██╔════╝██║░░░██║████╗░██║██╔══██╗██╔══██╗╚██╗░██╔╝
█████╗░░██║░░░██║██╔██╗██║██████╔╝███████║░╚████╔╝░
██╔══╝░░██║░░░██║██║╚████║██╔═══╝░██╔══██║░░╚██╔╝░░
██║░░░░░╚██████╔╝██║░╚███║██║░░░░░██║░░██║░░░██║░░░
╚═╝░░░░░░╚═════╝░╚═╝░░╚══╝╚═╝░░░░░╚═╝░░╚═╝░░░╚═╝░░░

░█████╗░██████╗░██████╗░███████╗██████╗░░██████╗
██╔══██╗██╔══██╗██╔══██╗██╔════╝██╔══██╗██╔════╝
██║░░██║██████╔╝██║░░██║█████╗░░██████╔╝╚█████╗░
██║░░██║██╔══██╗██║░░██║██╔══╝░░██╔══██╗░╚═══██╗
╚█████╔╝██║░░██║██████╔╝███████╗██║░░██║██████╔╝
░╚════╝░╚═╝░░╚═╝╚═════╝░╚══════╝╚═╝░░╚═╝╚═════╝░

██████╗░░█████╗░██████╗░░██████╗███████╗██████╗░
██╔══██╗██╔══██╗██╔══██╗██╔════╝██╔════╝██╔══██╗
██████╔╝███████║██████╔╝╚█████╗░█████╗░░██████╔╝
██╔═══╝░██╔══██║██╔══██╗░╚═══██╗██╔══╝░░██╔══██╗
██║░░░░░██║░░██║██║░░██║██████╔╝███████╗██║░░██║
╚═╝░░░░░╚═╝░░╚═╝╚═╝░░╚═╝╚═════╝░╚══════╝╚═╝░░╚═╝
"""

print(Fore.LIGHTWHITE_EX + text)

print("")
print(Fore.LIGHTMAGENTA_EX + "v0.0.1")
print("")
print(Fore.BLUE +"By AlexTills")

init()

def print_hyperlink(text, url):
    hyperlink = f"{Fore.BLUE}{Style.BRIGHT}{text}{Style.RESET_ALL}"
    print(f"\033]8;;{url}\033\\{hyperlink}\033]8;;\033\\")

print_hyperlink("Мой профиль на FunPay", "https://funpay.com/users/831315/")

print("==============================================================================")

options = webdriver.ChromeOptions()

options.add_argument("--window-size=100,100")
options.add_argument("--window-position=0,0")
driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=options)
driver.minimize_window()

workbook = Workbook()

headers = ["Дата", "Заказ", "Покупатель", "Цена", "Категория", "Описание"]

sh4 = workbook.active
sh4.title='Итоги'
sh4.merge_cells('B1:D1')
sh4['B1'].value = 'Статус'
sh4['B1'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
sh4['B1'].alignment = Alignment(horizontal='center', vertical='center')

# Добавление границ для объединенной ячейки
for cell in sh4['B1:D1'][0]:
    cell.border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))

# Закрыт
sh4['B2'].value = 'Закрыт'
sh4['B2'].fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
sh4['B2'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
sh4['B2'].alignment = Alignment(horizontal='center', vertical='center')

# Оплачен
sh4['C2'].value = 'Оплачен'
sh4['C2'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
sh4['C2'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
sh4['C2'].alignment = Alignment(horizontal='center', vertical='center')

# Возврат
sh4['D2'].value = 'Возврат'
sh4['D2'].fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
sh4['D2'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
sh4['D2'].alignment = Alignment(horizontal='center', vertical='center')

# Заполнение ячеек и установка цвета фона и границ
for row, text in enumerate(['Количество заказов:', 'Сумма заказов:', 'Самая высокая цена:', 'Средняя цена:'], start=3):
    sh4[f'A{row}'].value = text
    sh4[f'A{row}'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    sh4[f'A{row}'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
    sh4[f'A{row}'].alignment = Alignment(horizontal='right', vertical='center')
sh4.column_dimensions[get_column_letter(1)].width = 30
sh4.column_dimensions[get_column_letter(2)].width = 30
sh4.column_dimensions[get_column_letter(3)].width = 30
sh4.column_dimensions[get_column_letter(4)].width = 30

# Лист 1
sh1 = workbook.create_sheet(title='Закрыт')
sh1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
merged_cell = sh1.cell(row=1, column=1)
merged_cell.value = "Статус заказа: Закрыт"
merged_cell.alignment = Alignment(horizontal='center', vertical='center')
merged_cell.alignment = Alignment(horizontal='center', vertical='center')
sh1.append(headers)
for col_num in range(1, 6):
    cell = sh1.cell(row=2, column=col_num)
    cell.alignment = Alignment(horizontal='center', vertical='center')
sh1.column_dimensions[get_column_letter(1)].width = 21
sh1.column_dimensions[get_column_letter(2)].width = 13
sh1.column_dimensions[get_column_letter(3)].width = 14
sh1.column_dimensions[get_column_letter(4)].width = 10
sh1.column_dimensions[get_column_letter(5)].width = 36
sh1.column_dimensions[get_column_letter(6)].width = 20

# Лист 2
sh2 = workbook.create_sheet(title='Оплачен')
sh2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
merged_cell = sh2.cell(row=1, column=1)
merged_cell.value = "Статус заказа: Оплачен"
merged_cell.alignment = Alignment(horizontal='center', vertical='center')
merged_cell.alignment = Alignment(horizontal='center', vertical='center')
sh2.append(headers)
for col_num in range(1, 6):
    cell = sh2.cell(row=2, column=col_num)
    cell.alignment = Alignment(horizontal='center', vertical='center')
sh2.column_dimensions[get_column_letter(1)].width = 21
sh2.column_dimensions[get_column_letter(2)].width = 13
sh2.column_dimensions[get_column_letter(3)].width = 14
sh2.column_dimensions[get_column_letter(4)].width = 10
sh2.column_dimensions[get_column_letter(5)].width = 36
sh2.column_dimensions[get_column_letter(6)].width = 20

# Лист 3
sh3 = workbook.create_sheet(title='Вовзрат')
sh3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
merged_cell = sh3.cell(row=1, column=1)
merged_cell.value = "Статус заказа: Возврат"
merged_cell.alignment = Alignment(horizontal='center', vertical='center')
merged_cell.alignment = Alignment(horizontal='center', vertical='center')
sh3.append(headers)
for col_num in range(1, 6):
    cell = sh3.cell(row=2, column=col_num)
    cell.alignment = Alignment(horizontal='center', vertical='center')
sh3.column_dimensions[get_column_letter(1)].width = 21
sh3.column_dimensions[get_column_letter(2)].width = 13
sh3.column_dimensions[get_column_letter(3)].width = 14
sh3.column_dimensions[get_column_letter(4)].width = 10
sh3.column_dimensions[get_column_letter(5)].width = 36
sh3.column_dimensions[get_column_letter(6)].width = 20

price_column_index = 4
order_column_index = 2 

green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid") 
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  
orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  
grey_fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")  

for sheet in [sh1]:
    for cell in sheet[1]: 
        cell.fill = green_fill

for sheet in [sh2]:
    for cell in sheet[1]: 
        cell.fill = yellow_fill

for sheet in [sh3]:
    for cell in sheet[1]: 
        cell.fill = orange_fill

for sheet in [sh1, sh2, sh3]:
    for cell in sheet[2]: 
        cell.fill = grey_fill

border_style = Border(left=Side(style='thin'),
                      right=Side(style='thin'),
                      top=Side(style='thin'),
                      bottom=Side(style='thin'))

for sheet in workbook:
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = border_style

print("")

try:

    driver.get("https://funpay.com/orders/trade?id=&buyer=&state=closed&game=")

    try:
        with open(cookie_file_path, "r") as file:
            cookies = eval(file.read())
            for cookie in cookies:
                driver.add_cookie(cookie)
    except FileNotFoundError:
        raise FileNotFoundError("Файл cookies.txt не найден")
    except SyntaxError:
        raise SyntaxError(Fore.LIGHTRED_EX + "Файл cookies.txt не найден")
    
except Exception as e:
    print("Произошла ошибка:", e)



driver.refresh()

if driver.current_url == "https://funpay.com/account/login":
    print("Запустите файл login.bat и войдите в свой аккаунт FunPay, а затем нажмите Enter в консоли")
    driver.quit()
    exit()

time.sleep(2)

driver.get("https://funpay.com/orders/trade?id=&buyer=&state=paid&game=")

time.sleep(2)

print(Fore.GREEN + "Вход в аккаунт выполнен успешно")
print("")
print(Fore.LIGHTMAGENTA_EX + "Обработка заказов со статусом «Закрыт»...")
print("")

# 1 Цикл
while True:
    try:
        show_more_button = driver.find_element(By.CSS_SELECTOR, '.btn.btn-default.dyn-table-continue')
        show_more_button.click()
        time.sleep(0.3)

        if 'hidden' in show_more_button.get_attribute('class'):
            break

    except NoSuchElementException:
        break

time.sleep(2)

page_source = driver.page_source
soup = BeautifulSoup(page_source, 'html.parser')

orders1 = soup.find_all("a", class_="tc-item")

if orders1:
    prices1 = []
    buyers1 = []

    price_elements1 = soup.find_all(class_='tc-price text-nowrap tc-seller-sum')
    buyer_elements1 = soup.find_all("span", class_="pseudo-a")

    for price_element1 in price_elements1:
        price1 = price_element1.get_text().strip()
        prices1.append(price1)
        
    for buyer_element1 in buyer_elements1:
        buyer_name1 = buyer_element1.text.strip()
        buyer_url1 = buyer_element1['data-href']
        buyers1.append((buyer_name1, buyer_url1))

    for order1, price1 in zip(orders1, prices1):
        # Извлекаем информацию о заказе из элементов HTML
        date1 = order1.find("div", class_="tc-date-time").text.strip()
        order_num1 = order1.find("div", class_="tc-order").text.strip()
        description1 = order1.find("div", class_="order-desc").text.strip()
        category1 = order1.find("div", class_="text-muted").text.strip()

        sh1.append([date1, order_num1,"", price1, category1, description1])

    for index1, (buyer_name1, buyer_url1) in enumerate(buyers1, start=3):
        sh1.cell(row=index1, column=3).value = buyer_name1
        sh1.cell(row=index1, column=3).hyperlink = buyer_url1
        sh1.cell(row=index1, column=3).font = Font(underline="single", color="0563C1")

    total_price1 = 0
    for cell in sh1.iter_rows(min_row=3, min_col=price_column_index, max_col=price_column_index):
        for cell in cell:
            price = cell.value
            if price:
                price = float(price.replace(" ₽", "").replace(",", "").replace(" ", ""))
                price = round(price, 2)
                total_price1 += price

    max_price1 = None
    for cell in sh1.iter_rows(min_row=3, min_col=price_column_index, max_col=price_column_index):
        for cell in cell:
            price = cell.value
            if price:
                price = float(price.replace(" ₽", "").replace(",", "").replace(" ", ""))
                price = round(price, 2)
                if max_price1 is None or price > max_price1:
                    max_price1 = price

    filled_rows1 = sh1.max_row - 2
    average_price1 = total_price1 / filled_rows1
    average_price_str1 = "{:.2f}".format(average_price1)
    total_price_str1 = "{:.2f}".format(total_price1)
    max_price_str1 = "{:.2f}".format(max_price1)

    print(Fore.LIGHTGREEN_EX + "Статус «Закрыт»:")
    print(Fore.CYAN + "Количество Заказов:", Fore.LIGHTWHITE_EX, filled_rows1)
    print(Fore.CYAN + "Сумма Заказов:", Fore.LIGHTWHITE_EX, total_price_str1, "₽")
    print(Fore.CYAN + "Самая высокая цена", Fore.LIGHTWHITE_EX, max_price_str1, "₽")
    print(Fore.CYAN + "Средняя цена", Fore.LIGHTWHITE_EX, average_price_str1, "₽")

else:
    print(Fore.LIGHTRED_EX + "Заказы не найдены.")
    print("")


#2 цикл
driver.get("https://funpay.com/orders/trade?id=&buyer=&state=paid&game=")

time.sleep(2)

print("")
print(Fore.LIGHTMAGENTA_EX + "Обработка заказов со статусом «Оплачен»...")
print("")

while True:
    try:
        show_more_button = driver.find_element(By.CSS_SELECTOR, '.btn.btn-default.dyn-table-continue')
        show_more_button.click()
        time.sleep(0.3)

        if 'hidden' in show_more_button.get_attribute('class'):
            break

    except NoSuchElementException:
        break

time.sleep(3)

page_source = driver.page_source
soup = BeautifulSoup(page_source, 'html.parser')

orders2 = soup.find_all("a", class_="tc-item")


if orders2:

    prices2 = []
    buyers2 = []

    price_elements2 = soup.find_all(class_='tc-price text-nowrap tc-seller-sum')
    buyer_elements2 = soup.find_all("span", class_="pseudo-a")

    for price_element2 in price_elements2:
        price2 = price_element2.get_text().strip()
        prices2.append(price2)
        
    for buyer_element2 in buyer_elements2:
        buyer_name2 = buyer_element2.text.strip()
        buyer_url2 = buyer_element2['data-href']
        buyers2.append((buyer_name2, buyer_url2))

    for order2, price2 in zip(orders2, prices2):

        date2 = order2.find("div", class_="tc-date-time").text.strip()
        order_num2 = order2.find("div", class_="tc-order").text.strip()
        description2 = order2.find("div", class_="order-desc").text.strip()
        category2 = order2.find("div", class_="text-muted").text.strip()

        sh2.append([date2, order_num2,"", price2, category2, description2])

    for index2, (buyer_name2, buyer_url2) in enumerate(buyers2, start=3):
        sh2.cell(row=index2, column=3).value = buyer_name2
        sh2.cell(row=index2, column=3).hyperlink = buyer_url2
        sh2.cell(row=index2, column=3).font = Font(underline="single", color="0563C1")

    total_price2 = 0
    for cell in sh2.iter_rows(min_row=3, min_col=price_column_index, max_col=price_column_index):
        for cell in cell:
            price = cell.value
            if price:
                price = float(price.replace(" ₽", "").replace(",", "").replace(" ", ""))
                price = round(price, 2)
                total_price2 += price

    max_price2 = None
    for cell in sh2.iter_rows(min_row=3, min_col=price_column_index, max_col=price_column_index):
        for cell in cell:
            price = cell.value
            if price:
                price = float(price.replace(" ₽", "").replace(",", "").replace(" ", ""))
                price = round(price, 2)
                if max_price2 is None or price > max_price2:
                    max_price2 = price

    filled_rows2 = sh2.max_row - 2
    average_price2 = total_price2 / filled_rows2
    average_price_str2 = "{:.2f}".format(average_price2)
    total_price_str2 = "{:.2f}".format(total_price2)
    max_price_str2 = "{:.2f}".format(max_price2)

    print(Fore.LIGHTYELLOW_EX + "Статус «Оплачен»:")

    print(Fore.CYAN + "Количество Заказов:", Fore.LIGHTWHITE_EX, filled_rows2)
    print(Fore.CYAN + "Сумма Заказов:", Fore.LIGHTWHITE_EX, total_price_str2, "₽")
    print(Fore.CYAN + "Самая высокая цена:",  Fore.LIGHTWHITE_EX, max_price_str2, "₽")
    print(Fore.CYAN + "Средняя цена:", Fore.LIGHTWHITE_EX, average_price_str2, "₽")

else:
    print(Fore.LIGHTRED_EX + "Заказы не найдены.")
    print("")


#3 цикл
driver.get("https://funpay.com/orders/trade?id=&buyer=&state=refunded&game=")

time.sleep(2)

print("")
print(Fore.LIGHTMAGENTA_EX + "Обработка заказов со статусом «Возврат»...")
print("")

while True:
    try:
        show_more_button = driver.find_element(By.CSS_SELECTOR, '.btn.btn-default.dyn-table-continue')
        show_more_button.click()
        time.sleep(0.3)

        if 'hidden' in show_more_button.get_attribute('class'):
            break

    except NoSuchElementException:
        break

time.sleep(3)

page_source = driver.page_source
soup = BeautifulSoup(page_source, 'html.parser')

orders3 = soup.find_all("a", class_="tc-item")

if orders3:

    prices3 = []
    buyers3 = []

    price_elements3 = soup.find_all(class_='tc-price text-nowrap tc-seller-sum')
    buyer_elements3 = soup.find_all("span", class_="pseudo-a")

    for price_element3 in price_elements3:
        price3 = price_element3.get_text().strip()
        prices3.append(price3)

        
    for buyer_element3 in buyer_elements3:
        buyer_name3 = buyer_element3.text.strip()
        buyer_url3 = buyer_element3['data-href']
        buyers3.append((buyer_name3, buyer_url3))

    for order3, price3 in zip(orders3, prices3):
        date3 = order3.find("div", class_="tc-date-time").text.strip()
        order_num3 = order3.find("div", class_="tc-order").text.strip()
        description3 = order3.find("div", class_="order-desc").text.strip()
        category3 = order3.find("div", class_="text-muted").text.strip()

        sh3.append([date3, order_num3,"", price3, category3, description3])

    for index3, (buyer_name3, buyer_url3) in enumerate(buyers3, start=3):
        sh3.cell(row=index3, column=3).value = buyer_name3
        sh3.cell(row=index3, column=3).hyperlink = buyer_url3
        sh3.cell(row=index3, column=3).font = Font(underline="single", color="0563C1")

    total_price3 = 0
    for cell in sh3.iter_rows(min_row=3, min_col=price_column_index, max_col=price_column_index):
        for cell in cell:
            price = cell.value
            if price:
                price = float(price.replace(" ₽", "").replace(",", "").replace(" ", ""))
                price = round(price, 2)
                total_price3 += price

    max_price3 = None
    for cell in sh3.iter_rows(min_row=3, min_col=price_column_index, max_col=price_column_index):
        for cell in cell:
            price = cell.value
            if price:
                price = float(price.replace(" ₽", "").replace(",", "").replace(" ", ""))
                price = round(price, 2)
                if max_price3 is None or price > max_price3:
                    max_price3 = price

    filled_rows3 = sh3.max_row - 2
    average_price3 = total_price3 / filled_rows3
    average_price_str3 = "{:.2f}".format(average_price3)
    total_price_str3 = "{:.2f}".format(total_price3)
    max_price_str3 = "{:.2f}".format(max_price3)

    print(Fore.LIGHTRED_EX + "Статус «Возврат»:")

    print(Fore.CYAN + "Количество Заказов:", Fore.LIGHTWHITE_EX, filled_rows3)
    print(Fore.CYAN + "Сумма Заказов:", Fore.LIGHTWHITE_EX, total_price_str3, "₽")
    print(Fore.CYAN + "Самая высокая цена:", Fore.LIGHTWHITE_EX, max_price_str3, "₽")
    print(Fore.CYAN + "Средняя цена:", Fore.LIGHTWHITE_EX, average_price_str3, "₽")
else:
    print(Fore.LIGHTRED_EX + "Заказы не найдены.")
    print("")


for sh in [sh1, sh2, sh3]:
    for cell in sh.iter_rows(min_row=3, min_col=order_column_index, max_col=order_column_index):
        for cell in cell:
            order = cell.value
            if order:
                order = order.replace("#", "")
                url = f"https://funpay.com/orders/{order}/"
                cell.value = f"#{order}"
                cell.font = Font(underline="single", color="0563C1")
                cell.hyperlink = url


sh4['B3'] = filled_rows1
sh4['C3'] = filled_rows2
sh4['D3'] = filled_rows3

sh4['B4'] = total_price_str1 + " ₽"
sh4['C4'] = total_price_str2 + " ₽"
sh4['D4'] = total_price_str3 + " ₽"

sh4['B5'] = max_price_str1 + " ₽"
sh4['C5'] = max_price_str2 + " ₽"
sh4['D5'] = max_price_str3 + " ₽"

sh4['B6'] = average_price_str1 + " ₽"
sh4['C6'] = average_price_str2 + " ₽"
sh4['D6'] = average_price_str3 + " ₽"

for column in ['B', 'C', 'D']:
    for row in range(3, 7):
        cell = sh4[column + str(row)]
        cell.alignment = Alignment(horizontal='left', vertical='center')

time.sleep(1)

workbook.save("Results.xlsx")

print("")
print(Fore.LIGHTMAGENTA_EX + "Полная информация о заказах находится в файле Results.xlsx")
print("")